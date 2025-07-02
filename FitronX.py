from tkinter import Tk, Label, Entry, Radiobutton, Button, StringVar, IntVar, messagebox
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import zipfile
import random
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import pandas as pd
import time
import cv2
import mediapipe as mp
import numpy as np

# Dictionaries for prompts
#prompts
good_email = {
    "Congratulations on Your Outstanding Fitness Results!": """
    <html>
    <body>
    <p>Dear {Name},</p>

    <p>We are delighted to inform you about your exceptional performance in the recent fitness test conducted at NirveonX. Your results demonstrate a remarkable commitment to your health, and you should be incredibly proud of your accomplishments. 🎉</p>

    <p>The dedication you've invested in your fitness journey is truly commendable, and it's evident in your impressive scores. We're thrilled to announce that you're now prepared to <strong>elevate your training to the next level!</strong> 💪</p>

    <p>Here are the Nutritious diet to improve your Health based on Fitness:</p>

    <ul>
    <li><strong>Vegetarian:</strong></li>
    <ul>
    <li><strong>Breakfast:</strong> Oatmeal topped with fresh fruits (berries, banana) and a sprinkle of nuts and seeds. 🥣🍌</li>
    <li><strong>Mid-Morning Snack:</strong> Smoothie with kale, banana, almond butter, and almond milk. 🥤</li>
    <li><strong>Lunch:</strong> Brown rice with stir-fried tofu, mixed vegetables, and a soy-ginger sauce. 🍚🥦</li>
    <li><strong>Afternoon Snack:</strong> A piece of fruit (apple or orange) and some whole-grain crackers. 🍎🍊</li>
    <li><strong>Dinner:</strong> Stuffed bell peppers with quinoa, black beans, corn, and topped with cheese (optional). 🌶️🍽️</li>
    <li><strong>Beverages:</strong> Water, green tea, and occasional fresh fruit juices or smoothies. 💧🍵</li>
    </ul>

    <li><strong>Non-Vegetarian:</strong></li>
    <ul>
    <li><strong>Breakfast:</strong> Whole-grain toast with avocado and a poached egg. 🍞🥑🍳</li>
    <li><strong>Mid-Morning Snack:</strong> A handful of mixed nuts and a piece of fruit (apple or pear). 🥜🍏</li>
    <li><strong>Lunch:</strong> Turkey wrap with whole-grain tortilla, lettuce, tomatoes, and avocado. 🥙🥗</li>
    <li><strong>Afternoon Snack:</strong> Greek yogurt with honey and a few almonds. 🍯🍶</li>
    <li><strong>Dinner:</strong> Grilled chicken with brown rice and a side of steamed broccoli and carrots. 🍗🍚🥦🥕</li>
    <li><strong>Beverages:</strong> Water, herbal teas, and occasional fresh fruit juices or smoothies. 💧🍵</li>
    </ul>

    <p><strong>General Tips:</strong></p>
    <ul>
    <li><strong>Hydration:</strong> Drink plenty of water throughout the day. 💦</li>
    <li><strong>Avoid Processed Foods:</strong> Minimize intake of processed and sugary foods. 🚫🍭</li>
    <li><strong>Regular Meals:</strong> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
    <li><strong>Variety:</strong> Include a variety of foods to ensure a range of nutrients. 🌈</li>
    <li><strong>Cooking Methods:</strong> Prefer cooking methods like grilling, steaming, and baking over frying. 🔥</li>
    <li><strong>Portion Control:</strong> Be mindful of portion sizes, especially for high-calorie foods. ⚖️</li>
    </ul>

    <p>Are you eager to conquer your fitness goals? Our team of experienced trainers has compiled a selection of advanced programs designed to propel you towards achieving peak physical performance. 🌟</p>

    <p><strong>Fuel Your Body for Peak Performance:</strong> Imagine incorporating nutrient-rich, superhero-worthy greens into your diet, alongside protein powerhouses like eggs to facilitate muscle building. Don't forget to include a vibrant assortment of fruits and whole grains for sustained energy! 🥗🍳🍇</p>

    <p><strong>Advanced Workouts: Unleash Your Inner Athlete:</strong> Harness the expertise of our certified trainers who can design a personalized workout program specifically tailored to your needs. This program will keep your workouts fresh and engaging, pushing you beyond your current limits. 💪🏋️‍♂️</p>

    <p>To complement your new, challenging workouts, we can connect you with a registered dietitian to create a personalized nutrition plan that optimizes your fuel intake and supports your fitness dominance! 🍽️</p>

    <p>At NirveonX, we go beyond being just a gym; we're your dedicated fitness partner. We pledge to provide unwavering support throughout your fitness journey. Here's how you can maintain your winning streak:</p>

    <ul>
    <li><strong>Schedule a Follow-Up Consultation:</strong> Do you have questions about your results or your fitness aspirations? Let's discuss them together and design a program that empowers you to feel like an unstoppable fitness champion! 📅</li>
    <li><strong>Explore Advanced Fitness Programs:</strong> Discover a program that pushes your boundaries and empowers you to realize your full fitness potential! 🏆</li>
    </ul>

    <p>{Name}, you are a true source of inspiration! We are incredibly impressed with your progress and eagerly anticipate witnessing you conquer your next set of fitness goals! 🌟</p>

    <p>Stay strong and keep striving for greatness!</p>

    <p>Sincerely,</p>

    <p>NirveonX </p>

    <p>P.S. Have questions or ready to book a consultation? Just hit reply! We're always happy to help you become the best version of yourself. 💬</p>
    </body>
    </html>
    """,

    "Hey {Name}, High Five for Crushing Your Fitness Test!": """
    <html>
    <body>
    <p>Guess what? Your recent fitness test results at NirveonX are absolutely phenomenal! 🎉 You've built an incredible foundation for your health, and that's something to seriously brag about! 💪</p>

    <p>We know you've been putting in the hard work, and let us tell you, it's definitely paying off! Buckle up because here's the super exciting news: your results mean you're <strong>READY TO LEVEL UP!</strong> 🚀</p>

    <p>Feeling pumped to <strong>DOMINATE</strong> your fitness goals? We've got some epic suggestions to help you transform into an unstoppable fitness force! 🏆</p>

    <p>Here are the Nutritious diet to improve your Health based on Fitness:</p>

    <ul>
    <li><strong>Vegetarian:</strong></li>
    <ul>
    <li><strong>Breakfast:</strong> Oatmeal topped with fresh fruits (berries, banana) and a sprinkle of nuts and seeds. 🥣🍌</li>
    <li><strong>Mid-Morning Snack:</strong> Smoothie with kale, banana, almond butter, and almond milk. 🥤</li>
    <li><strong>Lunch:</strong> Brown rice with stir-fried tofu, mixed vegetables, and a soy-ginger sauce. 🍚🥦</li>
    <li><strong>Afternoon Snack:</strong> A piece of fruit (apple or orange) and some whole-grain crackers. 🍎🍊</li>
    <li><strong>Dinner:</strong> Stuffed bell peppers with quinoa, black beans, corn, and topped with cheese (optional). 🌶️🍽️</li>
    <li><strong>Beverages:</strong> Water, green tea, and occasional fresh fruit juices or smoothies. 💧🍵</li>
    </ul>

    <li><strong>Non-Vegetarian:</strong></li>
    <ul>
    <li><strong>Breakfast:</strong> Whole-grain toast with avocado and a poached egg. 🍞🥑🍳</li>
    <li><strong>Mid-Morning Snack:</strong> A handful of mixed nuts and a piece of fruit (apple or pear). 🥜🍏</li>
    <li><strong>Lunch:</strong> Turkey wrap with whole-grain tortilla, lettuce, tomatoes, and avocado. 🥙🥗</li>
    <li><strong>Afternoon Snack:</strong> Greek yogurt with honey and a few almonds. 🍯🍶</li>
    <li><strong>Dinner:</strong> Grilled chicken with brown rice and a side of steamed broccoli and carrots. 🍗🍚🥦🥕</li>
    <li><strong>Beverages:</strong> Water, herbal teas, and occasional fresh fruit juices or smoothies. 💧🍵</li>
    </ul>

    <p><strong>General Tips:</strong></p>
    <ul>
    <li><strong>Hydration:</strong> Drink plenty of water throughout the day. 💦</li>
    <li><strong>Avoid Processed Foods:</strong> Minimize intake of processed and sugary foods. 🚫🍭</li>
    <li><strong>Regular Meals:</strong> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
    <li><strong>Variety:</strong> Include a variety of foods to ensure a range of nutrients. 🌈</li>
    <li><strong>Cooking Methods:</strong> Prefer cooking methods like grilling, steaming, and baking over frying. 🔥</li>
    <li><strong>Portion Control:</strong> Be mindful of portion sizes, especially for high-calorie foods. ⚖️</li>
    </ul>

    <p>Ready to elevate your game? Our expert trainers are eager to design a personalized program to challenge you and take your fitness to new heights. 🚀</p>

    <p><strong>Power Up with Premium Nutrition:</strong> Imagine adding nutrient-dense greens and protein-packed eggs to your diet, providing the perfect fuel for muscle growth and energy. 🥗🍳🍇</p>

    <p><strong>Advanced Workouts: Challenge Accepted:</strong> Our certified trainers can craft a customized workout plan to keep your routine engaging and push your limits. 💪🏋️‍♂️</p>

    <p>Our nutrition experts are also ready to develop a personalized meal plan that aligns perfectly with your fitness goals. 🍽️</p>

    <p>At NirveonX, we're not just a gym; we're your dedicated fitness ally. Let's keep the momentum going!</p>

    <ul>
    <li><strong>Book a Follow-Up Consultation:</strong> Got questions or need guidance? We're here to help! 📅</li>
    <li><strong>Explore Advanced Fitness Programs:</strong> Discover programs that will challenge and inspire you! 🏆</li>
    </ul>

    <p>{Name}, your journey is truly inspiring, and we can't wait to see you crush your next set of goals. 🌟</p>

    <p>Keep pushing forward and stay amazing!</p>

    <p>Sincerely,</p>

    <p>NirveonX </p>

    <p>P.S. Ready to get started? Just reply to this email, and we’ll set everything up for you. 💬</p>
    </body>
    </html>
    """,

    "Your Fitness Journey Just Got More Exciting!": """
    <html>
    <body>
    <p>Hello {Name},</p>

    <p>We have some fantastic news to share with you! 🎉 Your recent fitness test results at NirveonX have been nothing short of outstanding. We're incredibly proud of your dedication and the progress you've made so far. 🌟</p>

    <p>Are you ready to take your fitness journey to the next level? With your impressive results, you're perfectly positioned to embark on a new adventure that will further enhance your well-being. 🚀</p>

    <p>Here are the Nutritious diet to improve your Health based on Fitness:</p>

    <ul>
    <li><strong>Vegetarian:</strong></li>
    <ul>
    <li><strong>Breakfast:</strong> Oatmeal topped with fresh fruits (berries, banana) and a sprinkle of nuts and seeds. 🥣🍌</li>
    <li><strong>Mid-Morning Snack:</strong> Smoothie with kale, banana, almond butter, and almond milk. 🥤</li>
    <li><strong>Lunch:</strong> Brown rice with stir-fried tofu, mixed vegetables, and a soy-ginger sauce. 🍚🥦</li>
    <li><strong>Afternoon Snack:</strong> A piece of fruit (apple or orange) and some whole-grain crackers. 🍎🍊</li>
    <li><strong>Dinner:</strong> Stuffed bell peppers with quinoa, black beans, corn, and topped with cheese (optional). 🌶️🍽️</li>
    <li><strong>Beverages:</strong> Water, green tea, and occasional fresh fruit juices or smoothies. 💧🍵</li>
    </ul>

    <li><strong>Non-Vegetarian:</strong></li>
    <ul>
    <li><strong>Breakfast:</strong> Whole-grain toast with avocado and a poached egg. 🍞🥑🍳</li>
    <li><strong>Mid-Morning Snack:</strong> A handful of mixed nuts and a piece of fruit (apple or pear). 🥜🍏</li>
    <li><strong>Lunch:</strong> Turkey wrap with whole-grain tortilla, lettuce, tomatoes, and avocado. 🥙🥗</li>
    <li><strong>Afternoon Snack:</strong> Greek yogurt with honey and a few almonds. 🍯🍶</li>
    <li><strong>Dinner:</strong> Grilled chicken with brown rice and a side of steamed broccoli and carrots. 🍗🍚🥦🥕</li>
    <li><strong>Beverages:</strong> Water, herbal teas, and occasional fresh fruit juices or smoothies. 💧🍵</li>
    </ul>

    <p><strong>General Tips:</strong></p>
    <ul>
    <li><strong>Hydration:</strong> Drink plenty of water throughout the day. 💦</li>
    <li><strong>Avoid Processed Foods:</strong> Minimize intake of processed and sugary foods. 🚫🍭</li>
    <li><strong>Regular Meals:</strong> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
    <li><strong>Variety:</strong> Include a variety of foods to ensure a range of nutrients. 🌈</li>
    <li><strong>Cooking Methods:</strong> Prefer cooking methods like grilling, steaming, and baking over frying. 🔥</li>
    <li><strong>Portion Control:</strong> Be mindful of portion sizes, especially for high-calorie foods. ⚖️</li>
    </ul>

    <p>To continue this incredible journey, we offer advanced training programs and nutritional consultations tailored just for you. 🌟</p>

    <p><strong>Optimize Your Nutrition:</strong> Incorporate nutrient-rich foods into your diet to fuel your body effectively. 🥗🍳🍇</p>

    <p><strong>Elevate Your Workouts:</strong> Engage with our trainers to design a dynamic workout plan that challenges you and helps you achieve your fitness goals. 💪🏋️‍♂️</p>

    <p>We can also connect you with a nutritionist to develop a personalized meal plan that complements your workout routine and maximizes your results. 🍽️</p>

    <p>At NirveonX, we're committed to supporting you every step of the way. Let's keep the momentum going!</p>

    <ul>
    <li><strong>Schedule a Consultation:</strong> Let's discuss your results and how we can help you achieve even greater success. 📅</li>
    <li><strong>Explore New Programs:</strong> Discover training programs that will keep you engaged and motivated! 🏆</li>
    </ul>

    <p>{Name}, your journey is nothing short of inspiring, and we can't wait to help you reach your next milestone. 🌟</p>

    <p>Keep up the amazing work!</p>

    <p>Sincerely,</p>

    <p>NirveonX </p>

    <p>P.S. Have any questions or ready to get started? Just reply to this email, and we'll take care of the rest. 💬</p>
    </body>
    </html>
    """,

    "Fitness Achievement Alert: You're Ready for the Next Step!": """
    <html>
    <body>
    <p>Dear {Name},</p>

    <p>Congratulations on your recent fitness test results at NirveonX! Your results are a testament to your hard work and dedication, and we couldn't be more excited for you. 🌟</p>

    <p>With your impressive performance, you're now poised to take your fitness journey to the next level. 🎉</p>

    <p>To support you in this next phase, we're thrilled to provide you with some nutritious diet options and expert tips to help you stay on track. 🏋️‍♂️💪</p>

    <p>Here are the Nutritious diet to improve your Health based on Fitness:</p>

    <ul>
    <li><strong>Vegetarian:</strong></li>
    <ul>
    <li><strong>Breakfast:</strong> Oatmeal topped with fresh fruits (berries, banana) and a sprinkle of nuts and seeds. 🥣🍌</li>
    <li><strong>Mid-Morning Snack:</strong> Smoothie with kale, banana, almond butter, and almond milk. 🥤</li>
    <li><strong>Lunch:</strong> Brown rice with stir-fried tofu, mixed vegetables, and a soy-ginger sauce. 🍚🥦</li>
    <li><strong>Afternoon Snack:</strong> A piece of fruit (apple or orange) and some whole-grain crackers. 🍎🍊</li>
    <li><strong>Dinner:</strong> Stuffed bell peppers with quinoa, black beans, corn, and topped with cheese (optional). 🌶️🍽️</li>
    <li><strong>Beverages:</strong> Water, green tea, and occasional fresh fruit juices or smoothies. 💧🍵</li>
    </ul>

    <li><strong>Non-Vegetarian:</strong></li>
    <ul>
    <li><strong>Breakfast:</strong> Whole-grain toast with avocado and a poached egg. 🍞🥑🍳</li>
    <li><strong>Mid-Morning Snack:</strong> A handful of mixed nuts and a piece of fruit (apple or pear). 🥜🍏</li>
    <li><strong>Lunch:</strong> Turkey wrap with whole-grain tortilla, lettuce, tomatoes, and avocado. 🥙🥗</li>
    <li><strong>Afternoon Snack:</strong> Greek yogurt with honey and a few almonds. 🍯🍶</li>
    <li><strong>Dinner:</strong> Grilled chicken with brown rice and a side of steamed broccoli and carrots. 🍗🍚🥦🥕</li>
    <li><strong>Beverages:</strong> Water, herbal teas, and occasional fresh fruit juices or smoothies. 💧🍵</li>
    </ul>

    <p><strong>General Tips:</strong></p>
    <ul>
    <li><strong>Hydration:</strong> Drink plenty of water throughout the day. 💦</li>
    <li><strong>Avoid Processed Foods:</strong> Minimize intake of processed and sugary foods. 🚫🍭</li>
    <li><strong>Regular Meals:</strong> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
    <li><strong>Variety:</strong> Include a variety of foods to ensure a range of nutrients. 🌈</li>
    <li><strong>Cooking Methods:</strong> Prefer cooking methods like grilling, steaming, and baking over frying. 🔥</li>
    <li><strong>Portion Control:</strong> Be mindful of portion sizes, especially for high-calorie foods. ⚖️</li>
    </ul>

    <p>We are excited to offer you personalized training programs that are designed to challenge and motivate you. 🚀</p>

    <p><strong>Elevate Your Nutrition:</strong> Boost your fitness journey with nutrient-dense foods and expert guidance. 🥗🍳🍇</p>

    <p><strong>Challenge Yourself with Advanced Workouts:</strong> Our trainers can create a dynamic and engaging workout plan just for you. 💪🏋️‍♂️</p>

    <p>Additionally, our nutritionists are available to develop a personalized meal plan that complements your fitness routine and maximizes your results. 🍽️</p>

    <p>At NirveonX, we're dedicated to supporting your fitness goals every step of the way. Let's keep the momentum going!</p>

    <ul>
    <li><strong>Book a Consultation:</strong> Discuss your results and next steps with our experts. 📅</li>
    <li><strong>Explore New Programs:</strong> Find the perfect program to push your limits and achieve your goals! 🏆</li>
    </ul>

    <p>{Name}, your progress is inspiring, and we can't wait to see you achieve your next milestone. 🌟</p>

    <p>Keep striving for greatness!</p>

    <p>Sincerely,</p>

    <p>NirveonX </p>

    <p>P.S. Have questions or ready to start? Just reply to this email, and we’ll help you get started. 💬</p>
    </body>
    </html>
    """,

    "Elevate Your Fitness Game with Our Personalized Programs!": """
    <html>
    <body>
    <p>Dear {Name},</p>

    <p>We are thrilled to share the amazing news about your recent fitness test results at NirveonX! Your dedication and hard work have truly paid off. 🌟</p>

    <p>With such outstanding results, you are now primed to take your fitness journey to new heights. 🚀</p>

    <p>To support you in this exciting phase, we're offering personalized training programs and nutrition plans designed just for you. 🏋️‍♂️💪</p>

    <p>Here are the Nutritious diet to improve your Health based on Fitness:</p>

    <ul>
    <li><strong>Vegetarian:</strong></li>
    <ul>
    <li><strong>Breakfast:</strong> Oatmeal topped with fresh fruits (berries, banana) and a sprinkle of nuts and seeds. 🥣🍌</li>
    <li><strong>Mid-Morning Snack:</strong> Smoothie with kale, banana, almond butter, and almond milk. 🥤</li>
    <li><strong>Lunch:</strong> Brown rice with stir-fried tofu, mixed vegetables, and a soy-ginger sauce. 🍚🥦</li>
    <li><strong>Afternoon Snack:</strong> A piece of fruit (apple or orange) and some whole-grain crackers. 🍎🍊</li>
    <li><strong>Dinner:</strong> Stuffed bell peppers with quinoa, black beans, corn, and topped with cheese (optional). 🌶️🍽️</li>
    <li><strong>Beverages:</strong> Water, green tea, and occasional fresh fruit juices or smoothies. 💧🍵</li>
    </ul>

    <li><strong>Non-Vegetarian:</strong></li>
    <ul>
    <li><strong>Breakfast:</strong> Whole-grain toast with avocado and a poached egg. 🍞🥑🍳</li>
    <li><strong>Mid-Morning Snack:</strong> A handful of mixed nuts and a piece of fruit (apple or pear). 🥜🍏</li>
    <li><strong>Lunch:</strong> Turkey wrap with whole-grain tortilla, lettuce, tomatoes, and avocado. 🥙🥗</li>
    <li><strong>Afternoon Snack:</strong> Greek yogurt with honey and a few almonds. 🍯🍶</li>
    <li><strong>Dinner:</strong> Grilled chicken with brown rice and a side of steamed broccoli and carrots. 🍗🍚🥦🥕</li>
    <li><strong>Beverages:</strong> Water, herbal teas, and occasional fresh fruit juices or smoothies. 💧🍵</li>
    </ul>

    <p><strong>General Tips:</strong></p>
    <ul>
    <li><strong>Hydration:</strong> Drink plenty of water throughout the day. 💦</li>
    <li><strong>Avoid Processed Foods:</strong> Minimize intake of processed and sugary foods. 🚫🍭</li>
    <li><strong>Regular Meals:</strong> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
    <li><strong>Variety:</strong> Include a variety of foods to ensure a range of nutrients. 🌈</li>
    <li><strong>Cooking Methods:</strong> Prefer cooking methods like grilling, steaming, and baking over frying. 🔥</li>
    <li><strong>Portion Control:</strong> Be mindful of portion sizes, especially for high-calorie foods. ⚖️</li>
    </ul>

    <p>Get ready to enhance your fitness journey with our advanced training programs and personalized nutrition plans. 🌟</p>

    <p><strong>Supercharge Your Nutrition:</strong> Incorporate nutrient-rich foods to fuel your workouts and recovery. 🥗🍳🍇</p>

    <p><strong>Challenge Yourself with Advanced Workouts:</strong> Our expert trainers can design a program that will keep you motivated and push your limits. 💪🏋️‍♂️</p>

    <p>We also offer personalized nutrition consultations to help you achieve your goals efficiently. 🍽️</p>

    <p>At NirveonX, we’re here to support your fitness journey every step of the way. Here’s how you can keep progressing:</p>

    <ul>
    <li><strong>Book a Follow-Up Consultation:</strong> Discuss your results and future goals with our experts. 📅</li>
    <li><strong>Explore Advanced Programs:</strong> Find the right program to help you achieve your fitness aspirations! 🏆</li>
    </ul>

    <p>{Name}, you're on an incredible path, and we’re excited to help you reach your next fitness milestone. 🌟</p>

    <p>Keep striving and stay awesome!</p>

    <p>Sincerely,</p>

    <p>NirveonX </p>

    <p>P.S. Have questions or ready to get started? Just reply to this email, and we’ll be happy to assist you. 💬</p>
    </body>
    </html>
    """
}

awesome_email = {
    f"Ready to Level Up Your Fitness, {{Name}}?": f"""<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; }}
        .bold {{ font-weight: bold; }}
    </style>
</head>
<body>
    <p>Hey {{Name}},</p>

    <p>Your recent fitness test scores are off the charts! 🎉 You've built an amazing foundation, and that's something to seriously celebrate! 🥳</p>

    <p><span class="bold">All your hard work is paying off</span>, and guess what? You're ready to take things to the <span class="bold">next level!</span> 💪 Feeling pumped?</p>

    <p><span class="bold">Continue the nutritious diet</span> you are following! But to improve your health more based on fitness, here are some nutritious diet options:</p>

    <p><span class="bold">Vegetarian:</span></p>
    <ul>
        <li><span class="bold">Breakfast:</span> Chia pudding made with almond milk, topped with fresh berries and a drizzle of honey. 🍓</li>
        <li><span class="bold">Mid-Morning Snack:</span> A green smoothie with spinach, kale, banana, and spirulina. 🥑</li>
        <li><span class="bold">Lunch:</span> Buddha bowl with brown rice, roasted chickpeas, sweet potatoes, spinach, avocado, and a tahini dressing. 🥗</li>
        <li><span class="bold">Afternoon Snack:</span> Mixed fruit bowl with seasonal fruits (mango, pineapple, berries). 🍍</li>
        <li><span class="bold">Dinner:</span> Vegetable stir-fry with tofu, broccoli, bell peppers, and snow peas, served over quinoa. 🍲</li>
        <li><span class="bold">Beverages:</span> Water, green tea, and fresh vegetable juices (carrot, beetroot). 🥤</li>
    </ul>

    <p><span class="bold">Non-Vegetarian:</span></p>
    <ul>
        <li><span class="bold">Breakfast:</span> Smoothie bowl with Greek yogurt, mixed berries, chia seeds, and a sprinkle of granola. 🍇</li>
        <li><span class="bold">Mid-Morning Snack:</span> A handful of raw almonds and a green apple. 🍎</li>
        <li><span class="bold">Lunch:</span> Grilled salmon salad with mixed greens, cherry tomatoes, cucumbers, and a lemon-dill dressing. 🐟</li>
        <li><span class="bold">Afternoon Snack:</span> Sliced bell peppers with hummus. 🌶️</li>
        <li><span class="bold">Dinner:</span> Grilled chicken breast with quinoa and a side of sautéed spinach and garlic. 🍗</li>
        <li><span class="bold">Beverages:</span> Water, herbal teas, and fresh vegetable juices (kale, cucumber, celery). 🍵</li>
    </ul>

    <p><span class="bold">General Tips:</span></p>
    <ul>
        <li><span class="bold">Hydration:</span> Drink plenty of water throughout the day. 💧</li>
        <li><span class="bold">Avoid Processed Foods:</span> Minimize intake of processed and sugary foods. 🍬</li>
        <li><span class="bold">Regular Meals:</span> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
        <li><span class="bold">Variety:</span> Include a variety of foods to ensure a range of nutrients. 🥦</li>
        <li><span class="bold">Cooking Methods:</span> Prefer cooking methods like grilling, steaming, and baking over frying. 🍳</li>
        <li><span class="bold">Portion Control:</span> Be mindful of portion sizes, especially for high-calorie foods. 📏</li>
    </ul>

    <p>We've got incredible plans designed to turn you into an <span class="bold">unstoppable fitness machine!</span> 🏋️ Our expert trainers will craft a personalized program that keeps things exciting and pushes you further than ever before.</p>

    <p>Think heavier weights, lightning-fast sprints, or even <span class="bold">mastering a brand new skill!</span> 🤸</p>

    <p>Want a personalized nutrition plan that fuels your fitness dominance? We can connect you with a registered dietitian to create a meal strategy that matches your rockstar workouts!</p>

    <p>Here at NirveonX, we're your ultimate fitness cheerleader, supporting you every step of the way.</p>

    <p>Let's keep this winning streak going! Schedule a follow-up consultation or explore our advanced programs designed to unleash your full potential.</p>

    <p>You're an inspiration, {{Name}}! We can't wait to see you crush your next goals. 💪</p>

    <p>Stay strong, stay awesome!</p>

    <p>Sincerely,</p>

    <p>NirveonX </p>

    <p>P.S. Have questions or ready to book a consultation? Just hit reply! We're always happy to help you become the best version of yourself.</p>
</body>
</html>""",

    f"Fitness Success! Time to Level Up Your Journey, {{Name}}!": f"""<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; }}
        .bold {{ font-weight: bold; }}
    </style>
</head>
<body>
    <p>Congratulations on your fitness success, {{Name}}! 🎉 Your recent results at NirveonX are amazing. You've built a strong foundation and are well on your fitness journey!</p>

    <p>We recognize your dedication, and here's the exciting news – you're <span class="bold">READY to LEVEL UP!</span> 🚀</p>

    <p><span class="bold">Continue the nutritious diet</span> you are following! But to improve your health more based on fitness, here are some nutritious diet options:</p>

    <p><span class="bold">Vegetarian:</span></p>
    <ul>
        <li><span class="bold">Breakfast:</span> Chia pudding made with almond milk, topped with fresh berries and a drizzle of honey. 🍓</li>
        <li><span class="bold">Mid-Morning Snack:</span> A green smoothie with spinach, kale, banana, and spirulina. 🥑</li>
        <li><span class="bold">Lunch:</span> Buddha bowl with brown rice, roasted chickpeas, sweet potatoes, spinach, avocado, and a tahini dressing. 🥗</li>
        <li><span class="bold">Afternoon Snack:</span> Mixed fruit bowl with seasonal fruits (mango, pineapple, berries). 🍍</li>
        <li><span class="bold">Dinner:</span> Vegetable stir-fry with tofu, broccoli, bell peppers, and snow peas, served over quinoa. 🍲</li>
        <li><span class="bold">Beverages:</span> Water, green tea, and fresh vegetable juices (carrot, beetroot). 🥤</li>
    </ul>

    <p><span class="bold">Non-Vegetarian:</span></p>
    <ul>
        <li><span class="bold">Breakfast:</span> Smoothie bowl with Greek yogurt, mixed berries, chia seeds, and a sprinkle of granola. 🍇</li>
        <li><span class="bold">Mid-Morning Snack:</span> A handful of raw almonds and a green apple. 🍎</li>
        <li><span class="bold">Lunch:</span> Grilled salmon salad with mixed greens, cherry tomatoes, cucumbers, and a lemon-dill dressing. 🐟</li>
        <li><span class="bold">Afternoon Snack:</span> Sliced bell peppers with hummus. 🌶️</li>
        <li><span class="bold">Dinner:</span> Grilled chicken breast with quinoa and a side of sautéed spinach and garlic. 🍗</li>
        <li><span class="bold">Beverages:</span> Water, herbal teas, and fresh vegetable juices (kale, cucumber, celery). 🍵</li>
    </ul>

    <p><span class="bold">General Tips:</span></p>
    <ul>
        <li><span class="bold">Hydration:</span> Drink plenty of water throughout the day. 💧</li>
        <li><span class="bold">Avoid Processed Foods:</span> Minimize intake of processed and sugary foods. 🍬</li>
        <li><span class="bold">Regular Meals:</span> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
        <li><span class="bold">Variety:</span> Include a variety of foods to ensure a range of nutrients. 🥦</li>
        <li><span class="bold">Cooking Methods:</span> Prefer cooking methods like grilling, steaming, and baking over frying. 🍳</li>
        <li><span class="bold">Portion Control:</span> Be mindful of portion sizes, especially for high-calorie foods. 📏</li>
    </ul>

    <p>Feeling the fire to achieve your fitness goals? 🔥 We've got customized plans to help you become an unstoppable force!</p>

    <p>Our expert trainers can help you design a fresh and exciting workout program that pushes you further than ever before.</p>

    <p>Think heavier weights, faster sprints, or even mastering a new skill to keep things engaging! 🏋️‍♂️</p>

    <p>Want a personalized nutrition plan to match your new training? We can connect you with a registered dietitian to create a meal strategy that fuels your journey to fitness dominance!</p>

    <p>We're your ultimate partner in fitness at NirveonX, here to support you every step of the way.</p>

    <p>Let's keep this momentum going! Schedule a follow-up consultation or explore our advanced programs to unlock your full fitness potential.</p>

    <p>You're an inspiration, {{Name}}! We can't wait to see you <span class="bold">conquer your next fitness milestones.</span> 💪</p>

    <p>Stay strong, stay awesome!</p>

    <p>Sincerely,</p>

    <p>NirveonX </p>

    <p>P.S. Have questions or ready to book a consultation? Just hit reply! We're always happy to help you become the best version of yourself.</p>
</body>
</html>""",

    f"Get Ready to Crush Your Fitness Goals, {{Name}}!": f"""<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; }}
        .bold {{ font-weight: bold; }}
    </style>
</head>
<body>
    <p>It's time to <span class="bold">CRUSH</span> your fitness goals, {{Name}}! 💥</p>

    <p>Your recent NirveonX fitness test results are impressive! 🎉 You've built a <span class="bold">Fantastic foundation</span>, and now it's time to leverage it!</p>

    <p><span class="bold">Continue the nutritious diet</span> you are following! But to improve your health more based on fitness, here are some nutritious diet options:</p>

    <p><span class="bold">Vegetarian:</span></p>
    <ul>
        <li><span class="bold">Breakfast:</span> Chia pudding made with almond milk, topped with fresh berries and a drizzle of honey. 🍓</li>
        <li><span class="bold">Mid-Morning Snack:</span> A green smoothie with spinach, kale, banana, and spirulina. 🥑</li>
        <li><span class="bold">Lunch:</span> Buddha bowl with brown rice, roasted chickpeas, sweet potatoes, spinach, avocado, and a tahini dressing. 🥗</li>
        <li><span class="bold">Afternoon Snack:</span> Mixed fruit bowl with seasonal fruits (mango, pineapple, berries). 🍍</li>
        <li><span class="bold">Dinner:</span> Vegetable stir-fry with tofu, broccoli, bell peppers, and snow peas, served over quinoa. 🍲</li>
        <li><span class="bold">Beverages:</span> Water, green tea, and fresh vegetable juices (carrot, beetroot). 🥤</li>
    </ul>

    <p><span class="bold">Non-Vegetarian:</span></p>
    <ul>
        <li><span class="bold">Breakfast:</span> Smoothie bowl with Greek yogurt, mixed berries, chia seeds, and a sprinkle of granola. 🍇</li>
        <li><span class="bold">Mid-Morning Snack:</span> A handful of raw almonds and a green apple. 🍎</li>
        <li><span class="bold">Lunch:</span> Grilled salmon salad with mixed greens, cherry tomatoes, cucumbers, and a lemon-dill dressing. 🐟</li>
        <li><span class="bold">Afternoon Snack:</span> Sliced bell peppers with hummus. 🌶️</li>
        <li><span class="bold">Dinner:</span> Grilled chicken breast with quinoa and a side of sautéed spinach and garlic. 🍗</li>
        <li><span class="bold">Beverages:</span> Water, herbal teas, and fresh vegetable juices (kale, cucumber, celery). 🍵</li>
    </ul>

    <p><span class="bold">General Tips:</span></p>
    <ul>
        <li><span class="bold">Hydration:</span> Drink plenty of water throughout the day. 💧</li>
        <li><span class="bold">Avoid Processed Foods:</span> Minimize intake of processed and sugary foods. 🍬</li>
        <li><span class="bold">Regular Meals:</span> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
        <li><span class="bold">Variety:</span> Include a variety of foods to ensure a range of nutrients. 🥦</li>
        <li><span class="bold">Cooking Methods:</span> Prefer cooking methods like grilling, steaming, and baking over frying. 🍳</li>
        <li><span class="bold">Portion Control:</span> Be mindful of portion sizes, especially for high-calorie foods. 📏</li>
    </ul>

    <p>Feeling motivated? We have just what you need to take things up a notch and become an <span class="bold">unstoppable fitness machine!</span> 🏋️‍♀️</p>

    <p>Our expert trainers will craft a personalized workout program designed to push you further than ever before.</p>

    <p>Think heavier weights, faster sprints, or even <span class="bold">mastering a new skill!</span> 🤸</p>

    <p>Want a personalized nutrition plan to match your workouts? We can connect you with a registered dietitian to create a meal strategy that fuels your fitness dominance!</p>

    <p>Ready to keep this winning streak alive? Schedule a follow-up consultation or explore our advanced programs designed to help you reach your <span class="bold">Full Potential.</span> 🌟</p>

    <p>You're an inspiration, {{Name}}! We're excited to see you <span class="bold">conquer your next fitness goals.</span> 💪</p>

    <p>Stay strong, stay awesome!</p>

    <p>Sincerely,</p>

    <p>NirveonX </p>

    <p>P.S. Have questions or ready to book a consultation? Just hit reply! We're always happy to help you become the best version of yourself.</p>
</body>
</html>""",

    f"Congratulations, {{Name}}! Unleash Your Fitness Potential!": f"""<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; }}
        .bold {{ font-weight: bold; }}
    </style>
</head>
<body>
    <p>Fantastic news, {{Name}}! 🎉 Your recent fitness test results are amazing! You've built a <span class="bold">fantastic foundation</span> for your health – something to be truly proud of!</p>

    <p>All your hard work is paying off, and we're thrilled to tell you – you're <span class="bold">ready to LEVEL UP!</span> 🚀</p>

    <p><span class="bold">Continue the nutritious diet</span> you are following! But to improve your health more based on fitness, here are some nutritious diet options:</p>

    <p><span class="bold">Vegetarian:</span></p>
    <ul>
        <li><span class="bold">Breakfast:</span> Chia pudding made with almond milk, topped with fresh berries and a drizzle of honey. 🍓</li>
        <li><span class="bold">Mid-Morning Snack:</span> A green smoothie with spinach, kale, banana, and spirulina. 🥑</li>
        <li><span class="bold">Lunch:</span> Buddha bowl with brown rice, roasted chickpeas, sweet potatoes, spinach, avocado, and a tahini dressing. 🥗</li>
        <li><span class="bold">Afternoon Snack:</span> Mixed fruit bowl with seasonal fruits (mango, pineapple, berries). 🍍</li>
        <li><span class="bold">Dinner:</span> Vegetable stir-fry with tofu, broccoli, bell peppers, and snow peas, served over quinoa. 🍲</li>
        <li><span class="bold">Beverages:</span> Water, green tea, and fresh vegetable juices (carrot, beetroot). 🥤</li>
    </ul>

    <p><span class="bold">Non-Vegetarian:</span></p>
    <ul>
        <li><span class="bold">Breakfast:</span> Smoothie bowl with Greek yogurt, mixed berries, chia seeds, and a sprinkle of granola. 🍇</li>
        <li><span class="bold">Mid-Morning Snack:</span> A handful of raw almonds and a green apple. 🍎</li>
        <li><span class="bold">Lunch:</span> Grilled salmon salad with mixed greens, cherry tomatoes, cucumbers, and a lemon-dill dressing. 🐟</li>
        <li><span class="bold">Afternoon Snack:</span> Sliced bell peppers with hummus. 🌶️</li>
        <li><span class="bold">Dinner:</span> Grilled chicken breast with quinoa and a side of sautéed spinach and garlic. 🍗</li>
        <li><span class="bold">Beverages:</span> Water, herbal teas, and fresh vegetable juices (kale, cucumber, celery). 🍵</li>
    </ul>

    <p><span class="bold">General Tips:</span></p>
    <ul>
        <li><span class="bold">Hydration:</span> Drink plenty of water throughout the day. 💧</li>
        <li><span class="bold">Avoid Processed Foods:</span> Minimize intake of processed and sugary foods. 🍬</li>
        <li><span class="bold">Regular Meals:</span> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
        <li><span class="bold">Variety:</span> Include a variety of foods to ensure a range of nutrients. 🥦</li>
        <li><span class="bold">Cooking Methods:</span> Prefer cooking methods like grilling, steaming, and baking over frying. 🍳</li>
        <li><span class="bold">Portion Control:</span> Be mindful of portion sizes, especially for high-calorie foods. 📏</li>
    </ul>

    <p>Ready to unleash your full fitness potential? 🔥 We have incredible programs designed to help you become an unstoppable force in your fitness journey!</p>

    <p>Our expert trainers are here to create a personalized program that keeps things exciting and pushes you to new heights.</p>

    <p>Think heavy lifting, fast sprints, or mastering new skills – whatever it takes to reach your goals! 🏆</p>

    <p>Want a personalized nutrition plan to complement your fitness program? We can connect you with a registered dietitian for a meal strategy that fuels your success!</p>

    <p>At NirveonX, we're here to support you every step of the way. Let’s keep this momentum going!</p>

    <p>We’re excited to see you <span class="bold">achieve your next big milestones.</span> 💪</p>

    <p>Stay strong, stay awesome!</p>

    <p>Sincerely,</p>

    <p>NirveonX </p>

    <p>P.S. Questions or ready to book a consultation? Just reply to this email! We’re always here to help you become the best version of yourself.</p>
</body>
</html>""",

    f"Congratulations on Your Achievement, {{Name}}! Let's Take the Next Step!": f"""<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; }}
        .bold {{ font-weight: bold; }}
    </style>
</head>
<body>
    <p>Congratulations on your recent fitness success, {{Name}}! 🎉 Your results are fantastic, and your progress is a testament to your dedication and hard work.</p>

    <p>It's time to <span class="bold">elevate</span> your fitness journey to new heights! 🚀</p>

    <p><span class="bold">Continue the nutritious diet</span> you are following! But to improve your health more based on fitness, here are some nutritious diet options:</p>

    <p><span class="bold">Vegetarian:</span></p>
    <ul>
        <li><span class="bold">Breakfast:</span> Chia pudding made with almond milk, topped with fresh berries and a drizzle of honey. 🍓</li>
        <li><span class="bold">Mid-Morning Snack:</span> A green smoothie with spinach, kale, banana, and spirulina. 🥑</li>
        <li><span class="bold">Lunch:</span> Buddha bowl with brown rice, roasted chickpeas, sweet potatoes, spinach, avocado, and a tahini dressing. 🥗</li>
        <li><span class="bold">Afternoon Snack:</span> Mixed fruit bowl with seasonal fruits (mango, pineapple, berries). 🍍</li>
        <li><span class="bold">Dinner:</span> Vegetable stir-fry with tofu, broccoli, bell peppers, and snow peas, served over quinoa. 🍲</li>
        <li><span class="bold">Beverages:</span> Water, green tea, and fresh vegetable juices (carrot, beetroot). 🥤</li>
    </ul>

    <p><span class="bold">Non-Vegetarian:</span></p>
    <ul>
        <li><span class="bold">Breakfast:</span> Smoothie bowl with Greek yogurt, mixed berries, chia seeds, and a sprinkle of granola. 🍇</li>
        <li><span class="bold">Mid-Morning Snack:</span> A handful of raw almonds and a green apple. 🍎</li>
        <li><span class="bold">Lunch:</span> Grilled salmon salad with mixed greens, cherry tomatoes, cucumbers, and a lemon-dill dressing. 🐟</li>
        <li><span class="bold">Afternoon Snack:</span> Sliced bell peppers with hummus. 🌶️</li>
        <li><span class="bold">Dinner:</span> Grilled chicken breast with quinoa and a side of sautéed spinach and garlic. 🍗</li>
        <li><span class="bold">Beverages:</span> Water, herbal teas, and fresh vegetable juices (kale, cucumber, celery). 🍵</li>
    </ul>

    <p><span class="bold">General Tips:</span></p>
    <ul>
        <li><span class="bold">Hydration:</span> Drink plenty of water throughout the day. 💧</li>
        <li><span class="bold">Avoid Processed Foods:</span> Minimize intake of processed and sugary foods. 🍬</li>
        <li><span class="bold">Regular Meals:</span> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
        <li><span class="bold">Variety:</span> Include a variety of foods to ensure a range of nutrients. 🥦</li>
        <li><span class="bold">Cooking Methods:</span> Prefer cooking methods like grilling, steaming, and baking over frying. 🍳</li>
        <li><span class="bold">Portion Control:</span> Be mindful of portion sizes, especially for high-calorie foods. 📏</li>
    </ul>

    <p>Ready to take things up a notch? 🔥 We offer amazing programs tailored to help you achieve your fitness goals and beyond!</p>

    <p>Our expert trainers will design a customized plan to challenge and motivate you every step of the way. Think of it as a thrilling new chapter in your fitness journey.</p>

    <p>Want to supercharge your progress? Connect with us to discuss a personalized nutrition plan or schedule a follow-up consultation to keep your momentum going.</p>

    <p>At NirveonX, we're dedicated to helping you reach your full potential.</p>

    <p>We’re thrilled to support you in <span class="bold">achieving your next big milestone!</span> 💪</p>

    <p>Stay strong, stay awesome!</p>

    <p>Sincerely,</p>

    <p>NirveonX </p>

    <p>P.S. Questions or ready to book a consultation? Just reply to this email! We're here to help you become the best version of yourself.</p>
</body>
</html>"""
}

poor_email = {
    f"Game On, {{Name}}! Your Fitness Journey Begins Here": f"""<p>Hi {{Name}}, high fives and victory dances for your recent fitness test! You're absolutely rocking it at NirveonX! 💪</p>

<p>Our results show you have the potential to be a <strong>Fitness Superstar!</strong> Ready to take it to the next level? We've got your back with expert trainers and registered dietitians to guide you. Think superhero-worthy greens to fuel your energy, protein to build muscle, and perfectly balanced meals to keep you strong.</p>

<p>Here are the nutritious diet options to improve your health based on fitness:</p>

<p><strong>Vegetarian:</strong></p>
<ul>
    <li><strong>Breakfast:</strong> Smoothie with spinach, banana, chia seeds, and almond milk. 🥑</li>
    <li><strong>Mid-Morning Snack:</strong> A handful of mixed nuts and a piece of fruit (apple or pear). 🍏</li>
    <li><strong>Lunch:</strong> Quinoa salad with chickpeas, cucumber, tomatoes, avocado, and a lemon-tahini dressing. 🥗</li>
    <li><strong>Afternoon Snack:</strong> Carrot and celery sticks with hummus. 🥕</li>
    <li><strong>Dinner:</strong> Lentil soup with a side of steamed vegetables (broccoli, carrots, and peas). 🍲</li>
    <li><strong>Beverages:</strong> Plenty of water, herbal teas, and fresh fruit juices. 🥤</li>
</ul>

<p><strong>Non-Vegetarian:</strong></p>
<ul>
    <li><strong>Breakfast:</strong> Greek yogurt with berries, honey, and a sprinkle of flax seeds. 🍇</li>
    <li><strong>Mid-Morning Snack:</strong> Hard-boiled eggs and a piece of fruit (orange or kiwi). 🍊</li>
    <li><strong>Lunch:</strong> Grilled chicken salad with mixed greens, cherry tomatoes, cucumbers, and a balsamic vinaigrette. 🥗</li>
    <li><strong>Afternoon Snack:</strong> Cottage cheese with pineapple chunks. 🍍</li>
    <li><strong>Dinner:</strong> Baked salmon with quinoa and a side of roasted vegetables (sweet potatoes, bell peppers, and asparagus). 🍣</li>
    <li><strong>Beverages:</strong> Plenty of water, green tea, and fresh fruit juices. 🥤</li>
</ul>

<p><strong>General Tips:</strong></p>
<ul>
    <li><strong>Hydration:</strong> Drink plenty of water throughout the day. 💧</li>
    <li><strong>Avoid Processed Foods:</strong> Minimize intake of processed and sugary foods. 🚫🍬</li>
    <li><strong>Regular Meals:</strong> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
    <li><strong>Variety:</strong> Include a variety of foods to ensure a range of nutrients. 🌈</li>
    <li><strong>Cooking Methods:</strong> Prefer cooking methods like grilling, steaming, and baking over frying. 🍳</li>
    <li><strong>Portion Control:</strong> Be mindful of portion sizes, especially for high-calorie foods. 📏</li>
</ul>

<p>Feeling overwhelmed? No worries! We'll tailor a workout plan that's challenging but totally achievable. Plus, we have beginner-friendly options to ease you in!</p>

<p>Here's how to keep your winning streak going: schedule a follow-up to create a program you love, and explore our awesome beginner-friendly classes.</p>

<p><strong>{{Name}}, you're incredible!</strong> We <strong>celebrate every victory</strong>, big or small, on your fitness journey. Let's make you the best version of yourself, together!</p>

<p>Stay strong, stay Healthy!</p>

<p>Sincerely,</p>

<p>NirveonX </p>

<p>P.S. Have questions or ready to book a consultation? Just hit reply! We're always happy to help you become the best version of yourself.</p>""",

    f"Calling All Champions! Your Fitness Journey Starts at NirveonX": f"""<p>Hi {{Name}}, high fives on your recent fitness test! You're rocking it at NirveonX, and we're so happy to have you on the team.</p>

<p>Our results show you have the potential to be an <strong>Absolute Beast</strong>. Ready to take it to the next level? We've got a whole crew of expert trainers and registered dietitians to guide you. Think energizing greens, muscle-building protein, and balanced meals to keep you powering through your workouts.</p>

<p>Here are the nutritious diet options to improve your health based on fitness:</p>

<p><strong>Vegetarian:</strong></p>
<ul>
    <li><strong>Breakfast:</strong> Smoothie with spinach, banana, chia seeds, and almond milk. 🥑</li>
    <li><strong>Mid-Morning Snack:</strong> A handful of mixed nuts and a piece of fruit (apple or pear). 🍏</li>
    <li><strong>Lunch:</strong> Quinoa salad with chickpeas, cucumber, tomatoes, avocado, and a lemon-tahini dressing. 🥗</li>
    <li><strong>Afternoon Snack:</strong> Carrot and celery sticks with hummus. 🥕</li>
    <li><strong>Dinner:</strong> Lentil soup with a side of steamed vegetables (broccoli, carrots, and peas). 🍲</li>
    <li><strong>Beverages:</strong> Plenty of water, herbal teas, and fresh fruit juices. 🥤</li>
</ul>

<p><strong>Non-Vegetarian:</strong></p>
<ul>
    <li><strong>Breakfast:</strong> Greek yogurt with berries, honey, and a sprinkle of flax seeds. 🍇</li>
    <li><strong>Mid-Morning Snack:</strong> Hard-boiled eggs and a piece of fruit (orange or kiwi). 🍊</li>
    <li><strong>Lunch:</strong> Grilled chicken salad with mixed greens, cherry tomatoes, cucumbers, and a balsamic vinaigrette. 🥗</li>
    <li><strong>Afternoon Snack:</strong> Cottage cheese with pineapple chunks. 🍍</li>
    <li><strong>Dinner:</strong> Baked salmon with quinoa and a side of roasted vegetables (sweet potatoes, bell peppers, and asparagus). 🍣</li>
    <li><strong>Beverages:</strong> Plenty of water, green tea, and fresh fruit juices. 🥤</li>
</ul>

<p><strong>General Tips:</strong></p>
<ul>
    <li><strong>Hydration:</strong> Drink plenty of water throughout the day. 💧</li>
    <li><strong>Avoid Processed Foods:</strong> Minimize intake of processed and sugary foods. 🚫🍬</li>
    <li><strong>Regular Meals:</strong> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
    <li><strong>Variety:</strong> Include a variety of foods to ensure a range of nutrients. 🌈</li>
    <li><strong>Cooking Methods:</strong> Prefer cooking methods like grilling, steaming, and baking over frying. 🍳</li>
    <li><strong>Portion Control:</strong> Be mindful of portion sizes, especially for high-calorie foods. 📏</li>
</ul>

<p>Feeling overwhelmed? No sweat! We'll customize a workout plan that's challenging but achievable. Plus, we have plenty of <strong>beginner-friendly options</strong> to get you started!</p>

<p>Here's how to keep your winning streak alive: schedule a follow-up to tailor a program you love, and explore beginner-friendly classes with your fellow <strong>Fitness Warriors</strong>.</p>

<p><strong>{{Name}}, you're incredible!</strong> We celebrate every step of your journey. Let's make you the best version of yourself, together!</p>

<p>Stay strong, stay Healthy!</p>

<p>Sincerely,</p>

<p>NirveonX </p>

<p>P.S. Have questions or ready to book a consultation? Just hit reply! We're always happy to help you become the best version of yourself.</p>""",

    f"Dive Deeper: Analyze Your Fitness Results and Level Up with NirveonX": f"""<p>Hey {{Name}},</p>

<p>We're thrilled with your recent fitness test results at NirveonX! Your results are a <strong>Great Starting Point</strong>, and based on the data, we see a <strong>Massive Potential</strong> for improvement.</p>

<p>Ready to transform into the ultimate athlete? We've got customized workout plans (lighter weights, gradual cardio progression, and focus on proper technique) designed by our expert trainers. Nutrition got you worried? No problem! We can connect you with a dietitian for personalized, delicious meals that support your specific goals.</p>

<p>Here are the nutritious diet options to improve your health based on fitness:</p>

<p><strong>Vegetarian:</strong></p>
<ul>
    <li><strong>Breakfast:</strong> Smoothie with spinach, banana, chia seeds, and almond milk. 🥑</li>
    <li><strong>Mid-Morning Snack:</strong> A handful of mixed nuts and a piece of fruit (apple or pear). 🍏</li>
    <li><strong>Lunch:</strong> Quinoa salad with chickpeas, cucumber, tomatoes, avocado, and a lemon-tahini dressing. 🥗</li>
    <li><strong>Afternoon Snack:</strong> Carrot and celery sticks with hummus. 🥕</li>
    <li><strong>Dinner:</strong> Lentil soup with a side of steamed vegetables (broccoli, carrots, and peas). 🍲</li>
    <li><strong>Beverages:</strong> Plenty of water, herbal teas, and fresh fruit juices. 🥤</li>
</ul>

<p><strong>Non-Vegetarian:</strong></p>
<ul>
    <li><strong>Breakfast:</strong> Greek yogurt with berries, honey, and a sprinkle of flax seeds. 🍇</li>
    <li><strong>Mid-Morning Snack:</strong> Hard-boiled eggs and a piece of fruit (orange or kiwi). 🍊</li>
    <li><strong>Lunch:</strong> Grilled chicken salad with mixed greens, cherry tomatoes, cucumbers, and a balsamic vinaigrette. 🥗</li>
    <li><strong>Afternoon Snack:</strong> Cottage cheese with pineapple chunks. 🍍</li>
    <li><strong>Dinner:</strong> Baked salmon with quinoa and a side of roasted vegetables (sweet potatoes, bell peppers, and asparagus). 🍣</li>
    <li><strong>Beverages:</strong> Plenty of water, green tea, and fresh fruit juices. 🥤</li>
</ul>

<p><strong>General Tips:</strong></p>
<ul>
    <li><strong>Hydration:</strong> Drink plenty of water throughout the day. 💧</li>
    <li><strong>Avoid Processed Foods:</strong> Minimize intake of processed and sugary foods. 🚫🍬</li>
    <li><strong>Regular Meals:</strong> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
    <li><strong>Variety:</strong> Include a variety of foods to ensure a range of nutrients. 🌈</li>
    <li><strong>Cooking Methods:</strong> Prefer cooking methods like grilling, steaming, and baking over frying. 🍳</li>
    <li><strong>Portion Control:</strong> Be mindful of portion sizes, especially for high-calorie foods. 📏</li>
</ul>

<p>Feeling lost in the data? Schedule a follow-up consultation! We'll analyze your results with you, answer your questions, and create a program that makes you feel confident and comfortable. Plus, we offer beginner-friendly classes to get you started on the right foot.</p>

<p><strong>{{Name}}, you're an inspiration!</strong> We can't wait to see you smash your fitness goals. Remember, it's a marathon, not a sprint, and we're here to celebrate every <strong>Milestone</strong> with you based on your progress!</p>

<p>Stay strong, stay Healthy!</p>

<p>Sincerely,</p>

<p>NirveonX </p>

<p>P.S. Have questions or ready to book a consultation? Just hit reply! We're always happy to help you become the best version of yourself.</p>""",

    f"You Got This! Level Up Your Fitness Game at NirveonX": f"""<p>Hey {{Name}},</p>

<p>🤜 on those fitness test results! We're seriously pumped to see you taking control of your health journey.</p>

<p>Looking to be the ultimate fitness machine? Our trainers can craft a personalized plan just for you (think manageable weights, gradually increasing cardio, and mastering the basics). You'll be feeling like a superhero in no time! 💥 Worried about fueling your body right? No sweat! We can connect you with a dietitian for easy, delicious meals that keep you energized and reaching your goals.</p>

<p>Here are the nutritious diet options to improve your health based on fitness:</p>

<p><strong>Vegetarian:</strong></p>
<ul>
    <li><strong>Breakfast:</strong> Smoothie with spinach, banana, chia seeds, and almond milk. 🥑</li>
    <li><strong>Mid-Morning Snack:</strong> A handful of mixed nuts and a piece of fruit (apple or pear). 🍏</li>
    <li><strong>Lunch:</strong> Quinoa salad with chickpeas, cucumber, tomatoes, avocado, and a lemon-tahini dressing. 🥗</li>
    <li><strong>Afternoon Snack:</strong> Carrot and celery sticks with hummus. 🥕</li>
    <li><strong>Dinner:</strong> Lentil soup with a side of steamed vegetables (broccoli, carrots, and peas). 🍲</li>
    <li><strong>Beverages:</strong> Plenty of water, herbal teas, and fresh fruit juices. 🥤</li>
</ul>

<p><strong>Non-Vegetarian:</strong></p>
<ul>
    <li><strong>Breakfast:</strong> Greek yogurt with berries, honey, and a sprinkle of flax seeds. 🍇</li>
    <li><strong>Mid-Morning Snack:</strong> Hard-boiled eggs and a piece of fruit (orange or kiwi). 🍊</li>
    <li><strong>Lunch:</strong> Grilled chicken salad with mixed greens, cherry tomatoes, cucumbers, and a balsamic vinaigrette. 🥗</li>
    <li><strong>Afternoon Snack:</strong> Cottage cheese with pineapple chunks. 🍍</li>
    <li><strong>Dinner:</strong> Baked salmon with quinoa and a side of roasted vegetables (sweet potatoes, bell peppers, and asparagus). 🍣</li>
    <li><strong>Beverages:</strong> Plenty of water, green tea, and fresh fruit juices. 🥤</li>
</ul>

<p><strong>General Tips:</strong></p>
<ul>
    <li><strong>Hydration:</strong> Drink plenty of water throughout the day. 💧</li>
    <li><strong>Avoid Processed Foods:</strong> Minimize intake of processed and sugary foods. 🚫🍬</li>
    <li><strong>Regular Meals:</strong> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
    <li><strong>Variety:</strong> Include a variety of foods to ensure a range of nutrients. 🌈</li>
    <li><strong>Cooking Methods:</strong> Prefer cooking methods like grilling, steaming, and baking over frying. 🍳</li>
    <li><strong>Portion Control:</strong> Be mindful of portion sizes, especially for high-calorie foods. 📏</li>
</ul>

<p>Feeling stuck? Schedule a follow-up with us! We're here to answer all your questions and design a program you absolutely love. We also have tons of <strong>beginner-friendly workouts</strong> to get you started.</p>

<p><strong>{{Name}}, you're an inspiration!</strong> We can't wait to see you absolutely crush those goals. Remember, it's a marathon, not a sprint. We're <strong>cheering you on every step</strong> of the way!</p>

<p>Stay strong, stay Healthy!</p>

<p>Sincerely,</p>

<p>NirveonX </p>

<p>P.S. Have questions or ready to book a consultation? Just hit reply! We're always happy to help you become the best version of yourself.</p>""",

    f"You Got This, {{Name}}! Level Up Your Fitness at NirveonX": f"""<p>Just wanted to give you a high five for absolutely crushing your recent fitness test at NirveonX! We're excited to see you taking charge of your health and building a strong foundation.</p>

<p>Looking to take your fitness to the next level and feel <strong>Stronger than Ever?</strong> Our expert trainers can create a personalized plan just for you. Think lighter weights to start, gradually increasing cardio, and mastering basic exercise techniques. You'll be amazed at how <strong>Quickly you Progress!</strong></p>

<p>Here are the nutritious diet options to improve your health based on fitness:</p>

<p><strong>Vegetarian:</strong></p>
<ul>
    <li><strong>Breakfast:</strong> Smoothie with spinach, banana, chia seeds, and almond milk. 🥑</li>
    <li><strong>Mid-Morning Snack:</strong> A handful of mixed nuts and a piece of fruit (apple or pear). 🍏</li>
    <li><strong>Lunch:</strong> Quinoa salad with chickpeas, cucumber, tomatoes, avocado, and a lemon-tahini dressing. 🥗</li>
    <li><strong>Afternoon Snack:</strong> Carrot and celery sticks with hummus. 🥕</li>
    <li><strong>Dinner:</strong> Lentil soup with a side of steamed vegetables (broccoli, carrots, and peas). 🍲</li>
    <li><strong>Beverages:</strong> Plenty of water, herbal teas, and fresh fruit juices. 🥤</li>
</ul>

<p><strong>Non-Vegetarian:</strong></p>
<ul>
    <li><strong>Breakfast:</strong> Greek yogurt with berries, honey, and a sprinkle of flax seeds. 🍇</li>
    <li><strong>Mid-Morning Snack:</strong> Hard-boiled eggs and a piece of fruit (orange or kiwi). 🍊</li>
    <li><strong>Lunch:</strong> Grilled chicken salad with mixed greens, cherry tomatoes, cucumbers, and a balsamic vinaigrette. 🥗</li>
    <li><strong>Afternoon Snack:</strong> Cottage cheese with pineapple chunks. 🍍</li>
    <li><strong>Dinner:</strong> Baked salmon with quinoa and a side of roasted vegetables (sweet potatoes, bell peppers, and asparagus). 🍣</li>
    <li><strong>Beverages:</strong> Plenty of water, green tea, and fresh fruit juices. 🥤</li>
</ul>

<p><strong>General Tips:</strong></p>
<ul>
    <li><strong>Hydration:</strong> Drink plenty of water throughout the day. 💧</li>
    <li><strong>Avoid Processed Foods:</strong> Minimize intake of processed and sugary foods. 🚫🍬</li>
    <li><strong>Regular Meals:</strong> Eat regular meals and snacks to maintain energy levels. 🍽️</li>
    <li><strong>Variety:</strong> Include a variety of foods to ensure a range of nutrients. 🌈</li>
    <li><strong>Cooking Methods:</strong> Prefer cooking methods like grilling, steaming, and baking over frying. 🍳</li>
    <li><strong>Portion Control:</strong> Be mindful of portion sizes, especially for high-calorie foods. 📏</li>
</ul>

<p>Feeling lost? Schedule a follow-up consultation with us! We'll review your results, answer your questions, and create a plan that works for you. Plus, we have <strong>beginner-friendly classes</strong> to help you ease into your new routine.</p>

<p><strong>{{Name}}, you're a rockstar!</strong> We're here to support you every step of the way. <strong>Enjoy every bit of your fitness journey</strong> and remember, <strong>we're cheering for you</strong>!</p>

<p>Stay strong, stay Healthy!</p>

<p>Sincerely,</p>

<p>NirveonX </p>

<p>P.S. Have questions or ready to book a consultation? Just hit reply! We're always happy to help you become the best version of yourself.</p>""",
}

# Randomly choose any key and value from the dictionaries
poor_sub, poor = random.choice(list(poor_email.items()))
good_sub, good = random.choice(list(good_email.items()))
awesome_sub, awesome = random.choice(list(awesome_email.items()))

# Encode email content
good_encoded = good.encode('utf-8')
awesome_encoded = awesome.encode('utf-8')
poor_encoded = poor.encode('utf-8')

goodsub_encoded = good_sub.encode('utf-8')
poorsub_encoded = poor_sub.encode('utf-8')
awesomesub_encoded = awesome_sub.encode('utf-8')

# Ideal data
#for bicep curls
bicep_curl_estimates_boys = {
    7: 15,
    8: 15,
    9: 15,
    10: 15,
    11: 15,
    12: 15,
    13: 25,
    14: 25,
    15: 35,
    16: 35,
    17: 35,
    18: 25,
    19: 25,
    20: 25,
    21: 25,
    22: 25,
    23: 25,
    24: 25,
    25: 25,
    26: 25,
    27: 25,
    28: 25,
    29: 25,
    30: 25,
    31: 25,
    32: 25,
    33: 25,
    34: 25,
    35: 25,
    36: 15,
    37: 15,
    38: 15,
    39: 15,
    40: 15,
    41: 15,
    42: 15,
    43: 15,
    44: 15,
    45: 15,
    46: 15,
    47: 15,
    48: 15,
    49: 15,
    50: 15
}

bicep_curl_estimates_girls = {
    7: 10,
    8: 10,
    9: 10,
    10: 10,
    11: 10,
    12: 10,
    13: 20,
    14: 20,
    15: 30,
    16: 30,
    17: 30,
    18: 20,
    19: 20,
    20: 20,
    21: 20,
    22: 20,
    23: 20,
    24: 20,
    25: 20,
    26: 20,
    27: 20,
    28: 20,
    29: 20,
    30: 20,
    31: 20,
    32: 20,
    33: 20,
    34: 20,
    35: 20,
    36: 10,
    37: 10,
    38: 10,
    39: 10,
    40: 10,
    41: 10,
    42: 10,
    43: 10,
    44: 10,
    45: 10,
    46: 10,
    47: 10,
    48: 10,
    49: 10,
    50: 10
}

bicep_curl_estimates_no_experience_boys = {
    7: 5,
    8: 5,
    9: 5,
    10: 5,
    11: 5,
    12: 5,
    13: 10,
    14: 10,
    15: 20,
    16: 20,
    17: 20,
    18: 10,
    19: 10,
    20: 10,
    21: 10,
    22: 10,
    23: 10,
    24: 10,
    25: 10,
    26: 10,
    27: 10,
    28: 10,
    29: 10,
    30: 10,
    31: 10,
    32: 10,
    33: 10,
    34: 10,
    35: 10,
    36: 5,
    37: 5,
    38: 5,
    39: 5,
    40: 5,
    41: 5,
    42: 5,
    43: 5,
    44: 5,
    45: 5,
    46: 5,
    47: 5,
    48: 5,
    49: 5,
    50: 5
}

bicep_curl_estimates_no_experience_girls = {
    7: 4,
    8: 4,
    9: 4,
    10: 4,
    11: 4,
    12: 4,
    13: 8,
    14: 8,
    15: 15,
    16: 15,
    17: 15,
    18: 8,
    19: 8,
    20: 8,
    21: 8,
    22: 8,
    23: 8,
    24: 8,
    25: 8,
    26: 8,
    27: 8,
    28: 8,
    29: 8,
    30: 8,
    31: 8,
    32: 8,
    33: 8,
    34: 8,
    35: 8,
    36: 4,
    37: 4,
    38: 4,
    39: 4,
    40: 4,
    41: 4,
    42: 4,
    43: 4,
    44: 4,
    45: 4,
    46: 4,
    47: 4,
    48: 4,
    49: 4,
    50: 4
}

#ideal data for squats
squat_estimates_boys = {
    7: 30,
    8: 30,
    9: 30,
    10: 30,
    11: 30,
    12: 30,
    13: 50,
    14: 50,
    15: 70,
    16: 70,
    17: 70,
    18: 50,
    19: 50,
    20: 50,
    21: 50,
    22: 50,
    23: 50,
    24: 50,
    25: 50,
    26: 50,
    27: 50,
    28: 50,
    29: 50,
    30: 50,
    31: 50,
    32: 50,
    33: 50,
    34: 50,
    35: 50,
    36: 30,
    37: 30,
    38: 30,
    39: 30,
    40: 30,
    41: 30,
    42: 30,
    43: 30,
    44: 30,
    45: 30,
    46: 30,
    47: 30,
    48: 30,
    49: 30,
    50: 30
}

squat_estimates_girls = {
    7: 25,
    8: 25,
    9: 25,
    10: 25,
    11: 25,
    12: 25,
    13: 40,
    14: 40,
    15: 60,
    16: 60,
    17: 60,
    18: 40,
    19: 40,
    20: 40,
    21: 40,
    22: 40,
    23: 40,
    24: 40,
    25: 40,
    26: 40,
    27: 40,
    28: 40,
    29: 40,
    30: 40,
    31: 40,
    32: 40,
    33: 40,
    34: 40,
    35: 40,
    36: 25,
    37: 25,
    38: 25,
    39: 25,
    40: 25,
    41: 25,
    42: 25,
    43: 25,
    44: 25,
    45: 25,
    46: 25,
    47: 25,
    48: 25,
    49: 25,
    50: 25
}

squat_estimates_no_experience_boys = {
    7: 15,
    8: 15,
    9: 15,
    10: 15,
    11: 15,
    12: 15,
    13: 25,
    14: 25,
    15: 40,
    16: 40,
    17: 40,
    18: 25,
    19: 25,
    20: 25,
    21: 25,
    22: 25,
    23: 25,
    24: 25,
    25: 25,
    26: 25,
    27: 25,
    28: 25,
    29: 25,
    30: 25,
    31: 25,
    32: 25,
    33: 25,
    34: 25,
    35: 25,
    36: 15,
    37: 15,
    38: 15,
    39: 15,
    40: 15,
    41: 15,
    42: 15,
    43: 15,
    44: 15,
    45: 15,
    46: 15,
    47: 15,
    48: 15,
    49: 15,
    50: 15
}

squat_estimates_no_experience_girls = {
    7: 10,
    8: 10,
    9: 10,
    10: 10,
    11: 10,
    12: 10,
    13: 20,
    14: 20,
    15: 30,
    16: 30,
    17: 30,
    18: 20,
    19: 20,
    20: 20,
    21: 20,
    22: 20,
    23: 20,
    24: 20,
    25: 20,
    26: 20,
    27: 20,
    28: 20,
    29: 20,
    30: 20,
    31: 20,
    32: 20,
    33: 20,
    34: 20,
    35: 20,
    36: 10,
    37: 10,
    38: 10,
    39: 10,
    40: 10,
    41: 10,
    42: 10,
    43: 10,
    44: 10,
    45: 10,
    46: 10,
    47: 10,
    48: 10,
    49: 10,
    50: 10
}

#ideal data for pushups
pushup_estimates_boys = {
    7: 25,
    8: 25,
    9: 25,
    10: 25,
    11: 25,
    12: 25,
    13: 45,
    14: 45,
    15: 60,
    16: 60,
    17: 60,
    18: 45,
    19: 45,
    20: 45,
    21: 45,
    22: 45,
    23: 45,
    24: 45,
    25: 45,
    26: 45,
    27: 45,
    28: 45,
    29: 45,
    30: 45,
    31: 45,
    32: 45,
    33: 45,
    34: 45,
    35: 45,
    36: 25,
    37: 25,
    38: 25,
    39: 25,
    40: 25,
    41: 25,
    42: 25,
    43: 25,
    44: 25,
    45: 25,
    46: 25,
    47: 25,
    48: 25,
    49: 25,
    50: 25
}

pushup_estimates_girls = {
    7: 20,
    8: 20,
    9: 20,
    10: 20,
    11: 20,
    12: 20,
    13: 35,
    14: 35,
    15: 45,
    16: 45,
    17: 45,
    18: 35,
    19: 35,
    20: 35,
    21: 35,
    22: 35,
    23: 35,
    24: 35,
    25: 35,
    26: 35,
    27: 35,
    28: 35,
    29: 35,
    30: 35,
    31: 35,
    32: 35,
    33: 35,
    34: 35,
    35: 35,
    36: 20,
    37: 20,
    38: 20,
    39: 20,
    40: 20,
    41: 20,
    42: 20,
    43: 20,
    44: 20,
    45: 20,
    46: 20,
    47: 20,
    48: 20,
    49: 20,
    50: 20
}

pushup_estimates_no_experience_boys = {
    7: 10,
    8: 10,
    9: 10,
    10: 10,
    11: 10,
    12: 10,
    13: 20,
    14: 20,
    15: 30,
    16: 30,
    17: 30,
    18: 20,
    19: 20,
    20: 20,
    21: 20,
    22: 20,
    23: 20,
    24: 20,
    25: 20,
    26: 20,
    27: 20,
    28: 20,
    29: 20,
    30: 20,
    31: 20,
    32: 20,
    33: 20,
    34: 20,
    35: 20,
    36: 10,
    37: 10,
    38: 10,
    39: 10,
    40: 10,
    41: 10,
    42: 10,
    43: 10,
    44: 10,
    45: 10,
    46: 10,
    47: 10,
    48: 10,
    49: 10,
    50: 10
}

pushup_estimates_no_experience_girls = {
    7: 8,
    8: 8,
    9: 8,
    10: 8,
    11: 8,
    12: 8,
    13: 15,
    14: 15,
    15: 25,
    16: 25,
    17: 25,
    18: 15,
    19: 15,
    20: 15,
    21: 15,
    22: 15,
    23: 15,
    24: 15,
    25: 15,
    26: 15,
    27: 15,
    28: 15,
    29: 15,
    30: 15,
    31: 15,
    32: 15,
    33: 15,
    34: 15,
    35: 15,
    36: 8,
    37: 8,
    38: 8,
    39: 8,
    40: 8,
    41: 8,
    42: 8,
    43: 8,
    44: 8,
    45: 8,
    46: 8,
    47: 8,
    48: 8,
    49: 8,
    50: 8
}

#ideal data for plank
plank_estimates_boys = {
    7: 30,
    8: 30,
    9: 30,
    10: 30,
    11: 30,
    12: 30,
    13: 45,
    14: 45,
    15: 60,
    16: 60,
    17: 60,
    18: 75,
    19: 75,
    20: 75,
    21: 75,
    22: 75,
    23: 75,
    24: 75,
    25: 75,
    26: 75,
    27: 75,
    28: 75,
    29: 75,
    30: 75,
    31: 75,
    32: 75,
    33: 75,
    34: 75,
    35: 75,
    36: 60,
    37: 60,
    38: 60,
    39: 60,
    40: 60,
    41: 60,
    42: 60,
    43: 60,
    44: 60,
    45: 60,
    46: 60,
    47: 60,
    48: 60,
    49: 60,
    50: 60
}

plank_estimates_girls = {
    7: 20,
    8: 20,
    9: 20,
    10: 20,
    11: 20,
    12: 20,
    13: 35,
    14: 35,
    15: 45,
    16: 45,
    17: 45,
    18: 60,
    19: 60,
    20: 60,
    21: 60,
    22: 60,
    23: 60,
    24: 60,
    25: 60,
    26: 60,
    27: 60,
    28: 60,
    29: 60,
    30: 60,
    31: 60,
    32: 60,
    33: 60,
    34: 60,
    35: 60,
    36: 45,
    37: 45,
    38: 45,
    39: 45,
    40: 45,
    41: 45,
    42: 45,
    43: 45,
    44: 45,
    45: 45,
    46: 45,
    47: 45,
    48: 45,
    49: 45,
    50: 45
}

plank_estimates_no_experience_boys = {
    7: 15,
    8: 15,
    9: 15,
    10: 15,
    11: 15,
    12: 15,
    13: 25,
    14: 25,
    15: 30,
    16: 30,
    17: 30,
    18: 45,
    19: 45,
    20: 45,
    21: 45,
    22: 45,
    23: 45,
    24: 45,
    25: 45,
    26: 45,
    27: 45,
    28: 45,
    29: 45,
    30: 45,
    31: 45,
    32: 45,
    33: 45,
    34: 45,
    35: 45,
    36: 30,
    37: 30,
    38: 30,
    39: 30,
    40: 30,
    41: 30,
    42: 30,
    43: 30,
    44: 30,
    45: 30,
    46: 30,
    47: 30,
    48: 30,
    49: 30,
    50: 30
}

plank_estimates_no_experience_girls = {
    7: 10,
    8: 10,
    9: 10,
    10: 10,
    11: 10,
    12: 10,
    13: 20,
    14: 20,
    15: 25,
    16: 25,
    17: 25,
    18: 35,
    19: 35,
    20: 35,
    21: 35,
    22: 35,
    23: 35,
    24: 35,
    25: 35,
    26: 35,
    27: 35,
    28: 35,
    29: 35,
    30: 35,
    31: 35,
    32: 35,
    33: 35,
    34: 35,
    35: 35,
    36: 25,
    37: 25,
    38: 25,
    39: 25,
    40: 25,
    41: 25,
    42: 25,
    43: 25,
    44: 25,
    45: 25,
    46: 25,
    47: 25,
    48: 25,
    49: 25,
    50: 25
}

# Function to send email
def send_email(sender_email, receiver_email, email_subject, email_body, rep_count):
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login("fitronx.nirveonx@gmail.com", "imaw pskl dxoe drtv")
        
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = receiver_email
        msg['Subject'] = email_subject
        msg.attach(MIMEText(email_body, 'html'))

        server.sendmail(sender_email, receiver_email, msg.as_string())
        server.quit()
        print("Email sent successfully.")
    except Exception as e:
        print("An error occurred:", e)

#daily adding a new sheet
def append_data_to_new_sheet(file_path, data, sheet_name):
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd
    import os
    import zipfile

    df = pd.DataFrame(data)
    print("DataFrame Columns:", df.columns)  # Debug print

    try:
        book = load_workbook(file_path)
    except FileNotFoundError:
        book = None
    except zipfile.BadZipFile:
        print(f"Error: The file '{file_path}' is not a valid Excel file or is corrupted.")
        return
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return

    if book is not None:
        # ✅ Check if sheet exists; if not, create it and write headers
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            sheet = book.create_sheet(sheet_name)
            sheet.append(list(df.columns))  # Add headers

        # ✅ Append new rows (excluding headers if sheet already exists)
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)

        book.save(file_path)

    else:
        # ✅ If file doesn't exist, create it with headers and data
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Data appended to sheet: {sheet_name}")
    print("DataFrame Columns:", df.columns)


def save_daily_data(file_path,data):
    print("Final Data Dictionary:", data)  # ✅ Debug: see what’s being saved
    sheet_name = datetime.now().strftime('%Y-%m-%d')
    append_data_to_new_sheet(file_path, data, sheet_name)
    messagebox.showinfo("Success", f"Data saved to sheet: {sheet_name}")



def reset_form():
    name_var.set("")
    age_var.set("")
    gender_var.set("0")
    email_var.set("")
    phone_number_var.set("")
    height_var.set("")
    weight_var.set("")
    workout_choice_var.set("0")
    experience_var.set("0")
    No_of_tries_var.set("")

#Function to declare a winner
extra = 0

def print_max_extra_rows(file_path, sheet_name):
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return
    except Exception as e:
        print(f"Error reading file: {e}")
        return

    # Remove leading and trailing spaces from column names
    df.columns = df.columns.str.strip()

    print("Columns in DataFrame:", df.columns)  # Confirm columns

    if 'Extra' not in df.columns:
        print("The 'Extra' column does not exist in the DataFrame.")
        return

    def get_max_extra_rows(group):
        max_extra = group['Extra'].max()
        return group[group['Extra'] == max_extra]

    max_rows = df.groupby('Gender').apply(get_max_extra_rows).reset_index(drop=True)

    results = ""
    for _, row in max_rows.iterrows():
        results += (f"Name: {row['Name']}\n"
                    f"Age: {row['Age']}\n"
                    f"Gender: {row['Gender']}\n"
                    f"No of tries: {row['No of tries']}\n"
                    f"Email: {row['Email']}\n"
                    "-----\n")

    # Display the results in a messagebox
    messagebox.showinfo("Winners", f"The Winners of the Fitness Test are: {results}")

def pick_winner():
    file_path = 'C:/Users/admin/OneDrive/Documents/workspace/projects/My Projects/Fitness project/FitronX Data.xlsx'
    sheet_name = datetime.now().strftime('%Y-%m-%d')
    print_max_extra_rows(file_path, sheet_name)

# Function to handle form submission
def submit_form():
    name = name_var.get()
    age = int(age_var.get())
    gender = gender_var.get()
    email = email_var.get()
    phone_number = phone_number_var.get()
    height = height_var.get()
    weight = weight_var.get()
    workout_choice = workout_choice_var.get()
    experience = experience_var.get()


    if not name or not email:
        messagebox.showerror("Error", "All fields are required.")
        return

    end_time = time.time() + 60 
    rep_count, plank_duration = track_workout(workout_choice, end_time,)  # Capture plank_duration

    if workout_choice != 4:
        No_of_tries_var.set(rep_count)
    else:
        No_of_tries_var.set(plank_duration) 

    # Compare rep count with ideal data
    extra = 0
    #bicep
    if workout_choice == 1:
        if experience == "s":
            if gender == "M":
                if rep_count > bicep_curl_estimates_boys[age] + 5:
                    extra = rep_count - bicep_curl_estimates_boys[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < bicep_curl_estimates_boys[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
            else:  # F
                if rep_count > bicep_curl_estimates_girls[age] + 5:
                    extra = rep_count - bicep_curl_estimates_girls[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < bicep_curl_estimates_girls[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
        else:  # no experience
            if gender == "M":
                if rep_count > bicep_curl_estimates_no_experience_boys[age] + 5:
                    extra = rep_count - bicep_curl_estimates_no_experience_boys[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < bicep_curl_estimates_no_experience_boys[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
            else:  # F
                if rep_count > bicep_curl_estimates_no_experience_girls[age] + 5:
                    extra = rep_count - bicep_curl_estimates_no_experience_girls[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < bicep_curl_estimates_no_experience_girls[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good

    # Squats
    elif workout_choice == 2:
        if experience == "s":
            if gender == "M":
                if rep_count > squat_estimates_boys[age] + 5:
                    extra = rep_count - squat_estimates_boys[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < squat_estimates_boys[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
            else:  # F
                if rep_count > squat_estimates_girls[age] + 5:
                    extra = rep_count - squat_estimates_girls[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < squat_estimates_girls[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
        else:  # no experience
            if gender == "M":
                if rep_count > squat_estimates_no_experience_boys[age] + 5:
                    extra = rep_count - squat_estimates_no_experience_boys[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < squat_estimates_no_experience_boys[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
            else:  # F
                if rep_count > squat_estimates_no_experience_girls[age] + 5:
                    extra = rep_count - squat_estimates_no_experience_girls[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < squat_estimates_no_experience_girls[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good

    # Pushups
    elif workout_choice == 3:
        if experience == "s":
            if gender == "M":
                if rep_count > pushup_estimates_boys[age] + 5:
                    extra = rep_count - pushup_estimates_boys[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < pushup_estimates_boys[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
            else:  # F
                if rep_count > pushup_estimates_girls[age] + 5:
                    extra = rep_count - pushup_estimates_girls[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < pushup_estimates_girls[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
        else:  # no experience
            if gender == "M":
                if rep_count > pushup_estimates_no_experience_boys[age] + 5:
                    extra = rep_count - pushup_estimates_no_experience_boys[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < pushup_estimates_no_experience_boys[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
            else:  # F
                if rep_count > pushup_estimates_no_experience_girls[age] + 5:
                    extra = rep_count - pushup_estimates_no_experience_girls[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif rep_count < pushup_estimates_no_experience_girls[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good

    # Plank
    elif workout_choice == 4:
        if experience == "s":
            if gender == "M":
                if plank_duration > plank_estimates_boys[age] + 5:
                    extra = plank_duration - plank_estimates_boys[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif plank_duration < plank_estimates_boys[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
            else:  # F
                if plank_duration > plank_estimates_girls[age] + 5:
                    extra = plank_duration - plank_estimates_girls[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif plank_duration < plank_estimates_girls[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
        else:  # no experience
            if gender == "M":
                if plank_duration > plank_estimates_no_experience_boys[age] + 5:
                    extra = plank_duration - plank_estimates_no_experience_boys[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif plank_duration < plank_estimates_no_experience_boys[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good
            else:  # F
                if plank_duration > plank_estimates_no_experience_girls[age] + 5:
                    extra = plank_duration - plank_estimates_no_experience_girls[age]
                    prompt_sub = awesome_sub
                    prompt_str = awesome
                elif plank_duration < plank_estimates_no_experience_girls[age] - 5:
                    prompt_sub = poor_sub
                    prompt_str = poor
                else:
                    prompt_sub = good_sub
                    prompt_str = good

    # Return results
    #return prompt_sub, prompt_str, extra

    # Format and send the email
    prompt_str = prompt_str.replace("{Name}", name)
    prompt_sub = prompt_sub.replace("{Name}", name)
    sender_email = "fitronx.nirveonx@gmail.com"
    receiver_email = email
    send_email(sender_email, receiver_email, prompt_sub, prompt_str,rep_count)

    # Save data to excel
    data = {
        'Name': [name_var.get()],
        'Age': [age_var.get()],
        'Gender': [gender_var.get()],
        'Email': [email_var.get()],
        'Phone Number': [phone_number_var.get()],
        'Height': [height_var.get()],
        'Weight': [weight_var.get()],
        'Workout Choice': [workout_choice_var.get()],    
        'No of tries': [rep_count],
        'Extra': [extra], 
        'Experience': [experience_var.get()],
        'Date': [pd.Timestamp.now()]
    }
    file_path = 'C:/Users/admin/OneDrive/Documents/workspace/projects/My Projects/Fitness project/FitronX Data.xlsx'
    save_daily_data(file_path,data)
    if workout_choice !=4:
        messagebox.showinfo("Success", f"Workout data saved successfully.\nTotal Reps: {rep_count}")
    else:
        messagebox.showinfo("Success", f"Workout data saved successfully.\nPlank Duration: {plank_duration}")
    reset_form()
#opencv
def calculate_angle(a, b, c):
    a = np.array(a)
    b = np.array(b)
    c = np.array(c)
    
    radians = np.arctan2(c[1] - b[1], c[0] - b[0]) - np.arctan2(a[1] - b[1], a[0] - b[0])
    angle = np.abs(radians * 180.0 / np.pi)
    if angle > 180.0:
        angle = 360 - angle
    return angle

def track_workout(workout_choice, end_time):
    cap = cv2.VideoCapture(0)
    rep_count = 0
    plank_duration = 0

    # Mouse callback function
    def close_camera(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            param['closed'] = True

    # Dictionary to keep track of the camera close status
    param = {'closed': False}

    # Create a window and set a mouse callback
    cv2.namedWindow('Workout Tracker')
    cv2.setMouseCallback('Workout Tracker', close_camera, param)

    if not cap.isOpened():
        print("Error: Unable to open camera.")
        return

    mp_drawing = mp.solutions.drawing_utils
    mp_pose = mp.solutions.pose

    with mp_pose.Pose(min_detection_confidence=0.5, min_tracking_confidence=0.5) as pose:
        last_command_time = {}
        workout_ended = False
        rep_count = 0
        last_command = None
        in_down_position = False
        plank_start_time = None
        plank_duration = 0
        
        while cap.isOpened() and not param['closed']:
            ret, frame = cap.read()
            if not ret:
                break

            image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            results = pose.process(image)
            landmarks = results.pose_landmarks

            if landmarks is not None:
                mp_drawing.draw_landmarks(image, landmarks, mp_pose.POSE_CONNECTIONS)

            try:
                landmarks = results.pose_landmarks.landmark

                # Your existing code with modifications
                if workout_choice == 1:  # Bicep Curls
                    shoulder_left = [landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].x,
                                    landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].y]
                    elbow_left = [landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value].y]
                    wrist_left = [landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value].x,
                                    landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value].y]

                    angle_left = calculate_angle(shoulder_left, elbow_left, wrist_left)

                    shoulder_right = [landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].x,
                                    landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].y]
                    elbow_right = [landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].x,
                                landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].y]
                    wrist_right = [landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value].x,
                                landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value].y]

                    angle_right = calculate_angle(shoulder_right, elbow_right, wrist_right)

                    if angle_right > 160 and angle_left > 160:
                        last_command = "DOWN"
                        last_command_time['down_bicep'] = time.time()
                        in_down_position = True
                    elif angle_right < 30 and angle_left < 30 and in_down_position:
                        last_command = "UP"
                        rep_count += 1
                        last_command_time['up_bicep'] = time.time()
                        in_down_position = False

                elif workout_choice == 2:  # Squats
                    knee_left = [landmarks[mp_pose.PoseLandmark.LEFT_KNEE.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_KNEE.value].y]
                    hip_left = [landmarks[mp_pose.PoseLandmark.LEFT_HIP.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_HIP.value].y]
                    ankle_left = [landmarks[mp_pose.PoseLandmark.LEFT_ANKLE.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_ANKLE.value].y]

                    angle_left = calculate_angle(hip_left, knee_left, ankle_left)

                    knee_right = [landmarks[mp_pose.PoseLandmark.RIGHT_KNEE.value].x,
                                landmarks[mp_pose.PoseLandmark.RIGHT_KNEE.value].y]
                    hip_right = [landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value].x,
                                landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value].y]
                    ankle_right = [landmarks[mp_pose.PoseLandmark.RIGHT_ANKLE.value].x,
                                landmarks[mp_pose.PoseLandmark.RIGHT_ANKLE.value].y]

                    angle_right = calculate_angle(hip_right, knee_right, ankle_right)

                    if angle_left < 60 and angle_right < 60 and not in_down_position:
                        last_command = "DOWN"
                        last_command_time['down_squats'] = time.time()
                        in_down_position = True
                    elif angle_left > 160 and angle_right > 160 and in_down_position:
                        last_command = "UP"
                        rep_count += 1
                        last_command_time['up_squats'] = time.time()
                        in_down_position = False
                  
                elif workout_choice == 3:  # Push-ups
                    # Left arm
                    shoulder_left = [landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].x,
                                    landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].y]
                    elbow_left = [landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value].y]
                    wrist_left = [landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value].y]

                    angle_left_arm = calculate_angle(shoulder_left, elbow_left, wrist_left)

                    # Right arm
                    shoulder_right = [landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].x,
                                    landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].y]
                    elbow_right = [landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].x,
                                landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].y]
                    wrist_right = [landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value].x,
                            landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value].y]

                    angle_right_arm = calculate_angle(shoulder_right, elbow_right, wrist_right)

                    # Left leg
                    knee_left = [landmarks[mp_pose.PoseLandmark.LEFT_KNEE.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_KNEE.value].y]
                    hip_left = [landmarks[mp_pose.PoseLandmark.LEFT_HIP.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_HIP.value].y]
                    ankle_left = [landmarks[mp_pose.PoseLandmark.LEFT_ANKLE.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_ANKLE.value].y]

                    angle_left_leg = calculate_angle(hip_left, knee_left, ankle_left)

                    # Right leg
                    knee_right = [landmarks[mp_pose.PoseLandmark.RIGHT_KNEE.value].x,
                                    landmarks[mp_pose.PoseLandmark.RIGHT_KNEE.value].y]
                    hip_right = [landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value].x,
                                    landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value].y]
                    ankle_right = [landmarks[mp_pose.PoseLandmark.RIGHT_ANKLE.value].x,
                                    landmarks[mp_pose.PoseLandmark.RIGHT_ANKLE.value].y]

                    angle_right_leg = calculate_angle(hip_right, knee_right, ankle_right)

                    if angle_left_arm > 90 and angle_right_arm > 90 and angle_left_leg > 160 and angle_right_leg > 160:
                        if not in_down_position:  # Ensure transition to "UP" position
                            last_command = "UP"
                            last_command_time['up_pushups'] = time.time()
                            in_down_position = False
                    elif angle_left_arm < 90 and angle_right_arm < 90:
                        if not in_down_position:  # Transition to "DOWN" position
                            last_command = "DOWN"
                            rep_count += 1
                            last_command_time['down_pushups'] = time.time()
                            in_down_position = True

                elif workout_choice == 4:  # Plank
                    # Left side
                    shoulder_left = [landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].x,
                                    landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].y]
                    elbow_left = [landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value].y]
                    wrist_left = [landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value].y]
                    hip_left = [landmarks[mp_pose.PoseLandmark.LEFT_HIP.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_HIP.value].y]
                    knee_left = [landmarks[mp_pose.PoseLandmark.LEFT_KNEE.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_KNEE.value].y]
                    ankle_left = [landmarks[mp_pose.PoseLandmark.LEFT_ANKLE.value].x,
                                landmarks[mp_pose.PoseLandmark.LEFT_ANKLE.value].y]

                    angle_left_arm = calculate_angle(shoulder_left, elbow_left, wrist_left)
                    angle_left_leg = calculate_angle(hip_left, knee_left, ankle_left)

                    # Right side
                    shoulder_right = [landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].x,
                                    landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].y]
                    elbow_right = [landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].x,
                                landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].y]
                    wrist_right = [landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value].x,
                                landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value].y]
                    hip_right = [landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value].x,
                                landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value].y]
                    knee_right = [landmarks[mp_pose.PoseLandmark.RIGHT_KNEE.value].x,
                                landmarks[mp_pose.PoseLandmark.RIGHT_KNEE.value].y]
                    ankle_right = [landmarks[mp_pose.PoseLandmark.RIGHT_ANKLE.value].x,
                                landmarks[mp_pose.PoseLandmark.RIGHT_ANKLE.value].y]

                    angle_right_arm = calculate_angle(shoulder_right, elbow_right, wrist_right)
                    angle_right_leg = calculate_angle(hip_right, knee_right, ankle_right)

                    if angle_left_arm > 80 and angle_left_arm < 100 and angle_right_arm > 80 and angle_right_arm < 100 and angle_left_leg > 160 and angle_right_leg > 160:
                        if plank_start_time is None:
                            plank_start_time = time.time()
                        plank_duration = time.time() - plank_start_time
                        last_command = "HOLD"
                    else:
                        last_command = "ADJUST"
                        plank_start_time = None
                        plank_duration = 0


                    print(f"Last Command: {last_command}, Plank Duration: {plank_duration:.2f}")
                    

                key = cv2.waitKey(1) & 0xFF
                if key == ord('e'):
                    workout_ended = True

                if workout_choice == 4:  # Display plank duration for planks
                    cv2.putText(image, f'Last Command: {last_command}', (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                    cv2.putText(image, f'Plank Duration: {plank_duration:.2f} s', (10, 90), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)
                else:
                    cv2.putText(image, f'Last Command: {last_command}', (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                    cv2.putText(image, f'Rep Count: {rep_count}', (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                

                image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
                cv2.imshow('Workout Tracker', image)

            except Exception as e:
                print(e)

            if time.time() > end_time or workout_ended:
                break    

        cap.release()
        cv2.destroyAllWindows()

        return rep_count, plank_duration


app = Tk()
app.title("Fitness Assessment Vision with AIML")
app.geometry("700x800") 
label_opts = {"fg": "black", "bg": "skyblue", "font": ("Arial", 14)}
entry_opts = {"font": ("Arial", 14)}

name_var = StringVar()
age_var = StringVar()
gender_var = IntVar()
email_var = StringVar()
phone_number_var = StringVar()
height_var = StringVar()
weight_var = StringVar()
workout_choice_var = IntVar()
No_of_tries_var = StringVar()
experience_var = IntVar()

heading_label = Label(app, text="FITNESS TEST FORM", font=("Helvetica", 16, "bold"))
heading_label.grid(row=0, column=0, columnspan=2, padx=10, pady=5)

subheading_label = Label(app, text="Know your Health status from your Fitness Standards", font=("Helvetica", 12))
subheading_label.grid(row=1, column=0, columnspan=2, padx=10, pady=5)

Label(app, text="Name:").grid(row=2, column=0, padx=10, pady=5)
Entry(app, textvariable=name_var).grid(row=2, column=1, padx=10, pady=5)

Label(app, text="Age:").grid(row=3, column=0, padx=10, pady=5)
Entry(app, textvariable= age_var).grid(row=3, column=1, padx=10, pady=5)

Label(app, text="Gender:").grid(row=4, column=0, padx=10, pady=5)
Radiobutton(app, text="Male", variable=gender_var, value=1).grid(row=4, column=1, padx=0, pady=5, sticky="w")
Radiobutton(app, text="Female", variable=gender_var, value=2).grid(row=4, column=2, padx=0, pady=5, sticky="w")

Label(app, text="Email:").grid(row=5, column=0, padx=10, pady=5)
Entry(app, textvariable= email_var).grid(row=5, column=1, padx=10, pady=5)

Label(app, text="Phone Number:").grid(row=6, column=0, padx=10, pady=5)
Entry(app, textvariable= phone_number_var).grid(row=6, column=1, padx=10, pady=5)

Label(app, text="Height:").grid(row=7, column=0, padx=10, pady=5)
Entry(app, textvariable= height_var).grid(row=7, column=1, padx=10, pady=5)

Label(app, text="Weight:").grid(row=8, column=0, padx=10, pady=5)
Entry(app, textvariable= weight_var).grid(row=8, column=1, padx=10, pady=5)

Label(app, text="Workout Choice:").grid(row=9, column=0, padx=10, pady=5)
Radiobutton(app, text="Bicep Curls", variable=workout_choice_var, value=1).grid(row=9, column=1, padx=10, pady=5, sticky="w")
Radiobutton(app, text="Squats", variable=workout_choice_var, value=2).grid(row=10, column=1, padx=10, pady=5, sticky="w")
Radiobutton(app, text="Pushups", variable=workout_choice_var, value=3).grid(row=11, column=1, padx=10, pady=5, sticky="w")
Radiobutton(app, text="Plank", variable=workout_choice_var, value=4).grid(row=12, column=1, padx=10, pady=5, sticky="w")

Label(app, text="Experience:").grid(row=13, column=0, padx=10, pady=5)
Radiobutton(app, text="Yes", variable=experience_var, value=1).grid(row=13, column=1, padx=0, pady=5, sticky="w")
Radiobutton(app, text="No", variable=experience_var, value=2).grid(row=13, column=2, padx=0, pady=5, sticky="w")

Button(app, text="Submit", command=submit_form).grid(row=14, column=0, columnspan=2, padx=10, pady=10)
Button(app, text="Pick Winner", command=pick_winner).grid(row=14, column=1, columnspan=2, padx=10, pady=10)

app.mainloop()